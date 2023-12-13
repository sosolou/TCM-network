import openpyxl
import csv


symbol =[]
symbollist = []
def get_data():
    excel = openpyxl.load_workbook("OMIM-Gene-Map-Retrieval.xlsx")
    sheet = excel['Sheet']
    rows = sheet.max_row
    for i in range(2, rows+1):
        Phenotype_str = sheet.cell(i, 10).value
        Approved_Symbol = sheet.cell(i, 6).value
        Phen1 = str(Phenotype_str)
        sub1 = 'Systemic lupus erythematosus'
        if sub1 in Phen1:
            symbol.append(Approved_Symbol)
            temp_list = symbol[:]
            symbollist.append(temp_list)
            del symbol[0]
        else:
            sub2 ="systemic lupus erythematosus"
            if sub2 in Phen1:
                symbol.append(Approved_Symbol)
                temp_list = symbol[:]
                symbollist.append(temp_list)
                del symbol[0]
            else:
                sub3 = "Immune"
                if sub3 in Phen1:
                    symbol.append(Approved_Symbol)
                    temp_list = symbol[:]
                    symbollist.append(temp_list)
                    del symbol[0]
                else:
                    sub4 ='immune'
                    if sub4 in Phen1:
                        symbol.append(Approved_Symbol)
                        temp_list = symbol[:]
                        symbollist.append(temp_list)
                        del symbol[0]
                    else:
                        Approved_Symbol = None
                        symbol.append(Approved_Symbol)
                        temp_list = symbol[:]
                        symbollist.append(temp_list)
                        del symbol[0]
    return symbollist

Symbol_list = []
writer_list = []
def page_spider():
    f = open('./OMIM.csv', mode='w', encoding='utf-8', newline='')
    writer = csv.writer(f)
    writer.writerow([ 'OMIM'])
    Symbol_list = get_data()
    print(Symbol_list)
    for detail in Symbol_list:
        writer_list.append(detail[0])
        writer.writerow(writer_list)
        del writer_list[0]
    f.close()


if __name__ == '__main__':
    page_spider()