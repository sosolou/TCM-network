import openpyxl
import csv 

symbol =[]
symbollist=[]
def get_data():
    excel = openpyxl.load_workbook("GeneCards-SearchResults.xlsx")
    sheet = excel['Sheet']
    rows = sheet.max_row
    for i in range(2, rows+1):
        Score = sheet.cell(i, 7).value
        Gene_Symbol = sheet.cell(i, 1).value
        if Score >= 1.0000:
           symbol.append(Gene_Symbol)
           templist = symbol[:]
           symbollist.append(templist)
           del symbol[0]
    return symbollist

Symbol_list = []
writerlist=[]
def page_spider():
    f = open('./Genecards.csv', mode='w', encoding='utf-8', newline='')
    writer = csv.writer(f)
    writer.writerow([ 'Gene cards'])
    Symbol_list = get_data()
    for detail in Symbol_list:
        writerlist.append(detail[0])
        writer.writerow(writerlist)
        del writerlist[0]
    f.close()


if __name__ == '__main__':
    page_spider()