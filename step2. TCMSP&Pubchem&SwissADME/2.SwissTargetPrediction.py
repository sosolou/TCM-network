import openpyxl
import time
from selenium.webdriver.common.by import By
from selenium.webdriver import Chrome
from selenium.webdriver.common.keys import Keys
from selenium.webdriver.chrome.options import Options
from selenium.webdriver.support import expected_conditions as ec
import csv


def login():
    url = 'http://www.swisstargetprediction.ch/'
    opt = Options()
    opt.page_load_strategy = 'eager'
    opt.add_experimental_option('excludeSwitches', ['enable-automation'])
    # opt.add_argument('--headless')
    web = Chrome(options=opt)
    web.get(url)
    return web


def is_exist_list(web):
    try:
        web.find_elements(By.XPATH, '//*[@id="resultTable"]/tbody/tr')
        return True
    except:
        return False


def is_exist(web):
    try:
        ec.alert_is_present()
        alert = web.switch_to.alert
        alert.accept()
        return True
    except:
        return False


def Searchpage(web):
    number = web.find_element(By.XPATH, '//*[@id="resultTable_info"] ').text
    number_str = number.split(',' )[5]
    if (int(number_str) % 15) == 0:
        pageNumber = int(number_str) / 15
    else:
        pageNumber = int(int(number_str) / 15) + 1
    return pageNumber


SwissTarget_list = []


def get_data():
    excel = openpyxl.load_workbook("SwissADME=1.xlsx")
    sheet = excel['SwissADME=1']
    rows = sheet.max_row
    for i in range(2, rows + 1):
        drug = sheet.cell(i, 1).value
        MOL_ID = sheet.cell(i, 2).value
        MOlecule = sheet.cell(i, 3).value
        SMILES = sheet.cell(i, 6).value
        web = login()
        time.sleep(6)
        web.find_element(By.XPATH, '//*[@id="smilesBox"]').send_keys(SMILES, Keys.ENTER)
        time.sleep(6)
        if is_exist(web):
            time.sleep(2)
            web.close()
        else:
            pages = Searchpage(web)
            for i in range(1, pages+1):
                td_list = web.find_elements(By.XPATH, '//*[@id="resultTable"]/tbody/tr')
                time.sleep(1)
                for td in td_list:
                    Probability = td.find_element(By.XPATH, './td[6]/span').text
                    if Probability > "0.05":
                        Target = td.find_element(By.XPATH, './td[1]').text
                        Common_name = td.find_element(By.XPATH, './td[2]/a').text
                        Uniprot_ID = td.find_element(By.XPATH, './td[3]/a').text
                        ChEMBL_ID = td.find_element(By.XPATH, './td[4]/a').text
                        Target_Class = td.find_element(By.XPATH, './td[5]').text
                        Probability = td.find_element(By.XPATH, './td[6]/span').text
                        SwissTarget_data = [drug, MOL_ID, MOlecule, SMILES, Target, Common_name, Uniprot_ID, ChEMBL_ID,
                                            Target_Class, Probability]
                        temp1_list = SwissTarget_data[:]
                        SwissTarget_list.append(temp1_list)
                web.find_element(By.XPATH, '//*[@id="resultTable_next"]').click()
                time.sleep(1)
            web.close()
            time.sleep(2)
    return SwissTarget_list


Swiss_list = []


def page_spider():
    f = open('./SwissTargetPrediction.csv', mode='w', encoding='utf-8', newline='')
    writer = csv.writer(f)
    writer.writerow(['drug', 'MOL_ID', 'MOlecule', 'SMILES', 'Target', 'Common_name', 'Uniprot_ID', 'ChEMBL_ID', 'Target_Class','Probability'])
    Swiss_list = get_data()
    writer_list = []
    for detail in Swiss_list:
        writer_list.append(detail[0])
        writer_list.append(detail[1])
        writer_list.append(detail[2])
        writer_list.append(detail[3])
        writer_list.append(detail[4])
        writer_list.append(detail[5])
        writer_list.append(detail[6])
        writer_list.append(detail[7])
        writer_list.append(detail[8])
        writer_list.append(detail[9])
        writer.writerow(writer_list)
        del writer_list[0:10]
    f.close()


if __name__ == '__main__':
    page_spider()
